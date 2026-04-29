import os
import re
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


IDENTITY_COLUMNS = [
    "Company Name",
    "Trading Code",
    "Scrip Code",
    "Sector",
    "Instrument Type",
    "Listing Year",
    "Market Category",
    "Electronic Share",
    "Debut Trading Date",
    "Present Operational Status",
    "Market Date",
    "Last Update",
]

PRICE_COLUMNS = [
    "LTP",
    "YCP",
    "Opening Price",
    "Adj Opening Price",
    "Closing Price",
    "Day Low",
    "Day High",
    "52W Low",
    "52W High",
    "Change Value",
    "Change %",
]

LIQUIDITY_COLUMNS = [
    "Day Trade No",
    "Day Volume",
    "Day Value (mn)",
]

MARKET_VALUE_COLUMNS = [
    "Market Cap (mn)",
    "Free Float Cap (mn)",
]

SHARE_CAPITAL_COLUMNS = [
    "Authorized Capital (mn)",
    "Paid-up Capital (mn)",
    "Total Securities",
    "Face Value",
    "Market Lot",
]

DEBT_AND_EQUITY_COLUMNS = [
    "Present Loan Status Date",
    "Short-term Loan (mn)",
    "Long-term Loan (mn)",
    "Total Loan (mn)",
    "Reserve & Surplus without OCI (mn)",
    "Other Comprehensive Income (OCI) (mn)",
]

EPS_PERIODS = [
    "Q1",
    "Q2",
    "HalfYearly",
    "Q3",
    "9Months",
    "Annual",
]

EPS_SUFFIXES = [
    "EPS",
    "EPS_COP",
    "Diluted_EPS_COP",
]

ANNUAL_PERFORMANCE_PREFIXES = [
    "Aud_EPS_COP_Basic",
    "Aud_EPS_COP_Diluted",
    "Aud_NAVPS",
    "Aud_PCO_mn",
    "Aud_Profit_mn",
    "Aud_TCI_mn",
]

PE_RATIO_SUFFIXES = [
    "PEwBasEPS",
    "PEwDilutEPS",
    "PEwTrailRatio",
    "PEwAuditBascEPS",
]

CORPORATE_ACTION_COLUMNS = [
    "Last Div Year",
    "Last Div Yield %",
    "Latest Dividend Status (%)",
    "Cash Dividend",
    "Bonus Issue (Stock Dividend)",
    "Right Issue",
    "Year End",
    "For the year ended",
    "Last AGM held on",
]

PROCESSED_BASE_COLUMNS = [
    "Company Name",
    "Trading Code",
    "Sector",
    "Market Category",
    "Present Operational Status",
    "LTP",
    "Market Cap (mn)",
    "Day Value (mn)",
    "Latest EPS Used",
    "Latest NAVPS Used",
    "Latest Profit Used (mn)",
    "Latest Operating Cash Flow Used (mn)",
    "Latest P/E Used",
    "Sector Avg P/E",
    "Sector Median P/E",
    "Sector Trimmed Avg P/E",
    "P/E vs Sector Median (%)",
    "Price to NAV",
    "Sector Avg P/NAV",
    "Sector Median P/NAV",
    "Sector Trimmed Avg P/NAV",
    "P/NAV vs Sector Median (%)",
    "Earnings Yield (%)",
    "Dividend Yield %",
    "Sector Avg Dividend Yield %",
    "Sector Median Dividend Yield %",
    "Dividend Yield vs Sector Median (%)",
    "Price Position 52W (%)",
    "Short-term Loan (mn)",
    "Long-term Loan (mn)",
    "Total Loan (mn)",
    "Debt to Market Cap",
    "Debt to Profit",
    "Cash Flow to Profit",
    "EPS Growth Available Years",
    "EPS Growth (%)",
    "Profit Growth Available Years",
    "Profit Growth (%)",
    "NAVPS Growth Available Years",
    "NAVPS Growth (%)",
    "Sponsor/Director Holding %",
    "Institute Holding %",
    "Foreign Holding %",
    "Public Holding %",
    "Sector Market Cap Rank",
    "Sector Liquidity Rank",
    "Sector Valid P/E Count",
    "Sector Outlier Count",
    "Usable For Sector Average?",
    "Outlier Flag",
    "Outlier Reason",
    "Metric Reliability Score",
    "Value Score",
    "Quality Score",
    "Risk Score",
    "Liquidity Score",
    "Dividend Score",
    "Final Screening Score",
    "Liquidity Signal",
    "Debt Risk Signal",
    "Dividend Signal",
    "Positive Signals",
    "Negative Signals",
    "Key Risks",
    "Primary Valuation Signal",
    "Primary Valuation Reason",
    "Watchlist Decision",
    "Watchlist Reason",
    "What To Check Next",
]

GUIDE_ROWS = [
    {
        "Section": "Purpose",
        "Item": "Workbook intent",
        "Meaning": "This workbook separates raw scraped DSE data from processed primary screening analysis.",
        "How To Use": "Use the processed sheets to shortlist companies for deeper research. Do not treat any label as an automatic buy or sell signal.",
    },
    {
        "Section": "Worksheets",
        "Item": "Raw_Scraped_Data",
        "Meaning": "Cleaned output from the scraper. Columns that are blank for every company are removed.",
        "How To Use": "Use this sheet when you need to verify the original scraped values behind a processed conclusion.",
    },
    {
        "Section": "Worksheets",
        "Item": "Processed_Analysis",
        "Meaning": "Calculated ratios, sector comparisons, scores, risks, reasons, and next-check guidance.",
        "How To Use": "Start here when screening companies. Company Name, Trading Code, and Sector are kept first for readability.",
    },
    {
        "Section": "Worksheets",
        "Item": "Sector_Summary",
        "Meaning": "Sector-level averages, medians, trimmed averages, outlier counts, and decision counts.",
        "How To Use": "Prefer sector median and trimmed average over simple average when outliers are present.",
    },
    {
        "Section": "Worksheets",
        "Item": "Watchlist",
        "Meaning": "A focused subset of companies that deserve follow-up review based on the screening model.",
        "How To Use": "Use this as a research queue, not as a purchase list.",
    },
    {
        "Section": "Worksheets",
        "Item": "Data_Quality_Issues",
        "Meaning": "Companies with missing, unreliable, unusual, or outlier data.",
        "How To Use": "Check this sheet before trusting a valuation label for a specific company.",
    },
    {
        "Section": "Worksheets",
        "Item": "Dropped_Empty_Columns",
        "Meaning": "Raw columns removed because every scraped company had a blank value.",
        "How To Use": "Use this sheet as an audit trail for why the raw worksheet looks cleaner.",
    },
    {
        "Section": "Valuation Labels",
        "Item": "Potentially Undervalued",
        "Meaning": "The company appears attractive on primary valuation measures while quality and risk checks are acceptable.",
        "How To Use": "Read the annual report, quarterly report, PSI disclosures, dividend history, debt notes, and peer comparison before considering purchase.",
    },
    {
        "Section": "Valuation Labels",
        "Item": "Value Trap Risk",
        "Meaning": "Some valuation metrics look cheap, but risk indicators weaken the case.",
        "How To Use": "Focus on debt, cash-flow conversion, falling profit, abnormal earnings, weak liquidity, or governance concerns.",
    },
    {
        "Section": "Valuation Labels",
        "Item": "Potentially Overvalued",
        "Meaning": "The stock looks expensive on primary valuation or price-position measures.",
        "How To Use": "Check whether growth, margins, brand strength, market leadership, or sector outlook justifies the premium.",
    },
    {
        "Section": "Valuation Labels",
        "Item": "Speculative / Risky",
        "Meaning": "Earnings, profit, debt, liquidity, or data quality is too weak for a clean valuation signal.",
        "How To Use": "Treat as high caution until the reason for weakness is understood from filings and disclosures.",
    },
    {
        "Section": "Valuation Labels",
        "Item": "Insufficient Data",
        "Meaning": "The scraper did not collect enough reliable data for a meaningful primary conclusion.",
        "How To Use": "Verify missing values manually from DSE, annual reports, and latest quarterly statements.",
    },
    {
        "Section": "Important Metrics",
        "Item": "P/E vs Sector Median (%)",
        "Meaning": "Shows how the company's P/E compares with the sector median. Negative means cheaper than sector median.",
        "How To Use": "A low value is useful only when earnings are recurring and not distorted by one-time gains.",
    },
    {
        "Section": "Important Metrics",
        "Item": "Price to NAV",
        "Meaning": "Compares market price with net asset value per share.",
        "How To Use": "Low P/NAV can be interesting for asset-heavy businesses, but asset quality and ROE must be checked.",
    },
    {
        "Section": "Important Metrics",
        "Item": "Cash Flow to Profit",
        "Meaning": "Compares operating cash flow with reported profit.",
        "How To Use": "Weak or negative cash conversion can mean profit quality needs deeper review.",
    },
    {
        "Section": "Important Metrics",
        "Item": "Debt to Market Cap",
        "Meaning": "Compares total loans with market capitalization.",
        "How To Use": "High values can signal balance-sheet pressure, especially when profit or cash flow is weak.",
    },
    {
        "Section": "Important Metrics",
        "Item": "Price Position 52W (%)",
        "Meaning": "Shows where the current price sits between the 52-week low and high.",
        "How To Use": "A low value may signal opportunity or trouble. A high value may reduce margin of safety.",
    },
    {
        "Section": "Follow-up",
        "Item": "What To Check Next",
        "Meaning": "Plain-language due-diligence prompts generated from each company's signals and risks.",
        "How To Use": "Use this as the checklist before making any purchase-consideration decision.",
    },
]


def existing_columns(columns, preferred_order):
    """Return preferred columns that exist in the scraped data."""
    return [column for column in preferred_order if column in columns]


def move_columns_to_front(df, front_columns):
    """Return a DataFrame with important identifier columns first."""
    front = existing_columns(df.columns, front_columns)
    remaining = [column for column in df.columns if column not in front]
    return df.reindex(columns=front + remaining)


def split_all_empty_columns(df):
    """Drop fully empty raw columns and return a small audit table."""
    protected_columns = {"Company Name", "Trading Code", "Scrip Code", "Sector"}
    normalized = df.replace(r"^\s*$", pd.NA, regex=True)
    empty_columns = [
        column
        for column in normalized.columns
        if column not in protected_columns and normalized[column].isna().all()
    ]

    cleaned_df = df.drop(columns=empty_columns)
    dropped_columns_df = pd.DataFrame(
        {
            "Dropped Empty Column": empty_columns,
            "Reason": ["Column was blank for every scraped company"]
            * len(empty_columns),
        }
    )
    return cleaned_df, dropped_columns_df


def is_annual_performance_column(column):
    """Return True for audited annual performance columns."""
    return any(
        column.startswith(f"{prefix}_") for prefix in ANNUAL_PERFORMANCE_PREFIXES
    )


def annual_performance_sort_key(column):
    """Sort annual performance fields by year, then by metric."""
    metric_order = {
        prefix: index for index, prefix in enumerate(ANNUAL_PERFORMANCE_PREFIXES)
    }

    for prefix in ANNUAL_PERFORMANCE_PREFIXES:
        column_prefix = f"{prefix}_"
        if column.startswith(column_prefix):
            year_part = column[len(column_prefix) :]
            year_tokens = [
                int(token)
                for token in year_part.replace("/", "_").replace("-", "_").split("_")
                if token.isdigit()
            ]
            sort_year = max(year_tokens) if year_tokens else 0
            return (-sort_year, metric_order[prefix], column)

    return (0, len(ANNUAL_PERFORMANCE_PREFIXES), column)


def is_eps_metric_column(column):
    """Return True for quarterly/annual EPS columns from the EPS table."""
    return any(
        column == f"{period}_{suffix}"
        for period in EPS_PERIODS
        for suffix in EPS_SUFFIXES
    )


def eps_metric_sort_key(column):
    """Sort EPS columns by period first, then EPS type."""
    for period_index, period in enumerate(EPS_PERIODS):
        for suffix_index, suffix in enumerate(EPS_SUFFIXES):
            if column == f"{period}_{suffix}":
                return (period_index, suffix_index, column)

    return (len(EPS_PERIODS), len(EPS_SUFFIXES), column)


def is_pe_ratio_column(column):
    """Return True for dynamic P/E ratio columns."""
    return any(column.endswith(f"_{suffix}") for suffix in PE_RATIO_SUFFIXES)


def parse_date_from_column(column):
    """Parse DSE date fragments that appear before a dynamic P/E suffix."""
    date_part = column
    for suffix in PE_RATIO_SUFFIXES:
        marker = f"_{suffix}"
        if column.endswith(marker):
            date_part = column[: -len(marker)]
            break

    cleaned = date_part.replace("_", " ").replace(",", "").strip()
    for fmt in ["%b %d %Y", "%B %d %Y"]:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            pass
    return datetime.min


def pe_ratio_sort_key(column):
    """Sort P/E ratio columns by metric family, then latest source date first."""
    suffix_order = {
        "PEwTrailRatio": 0,
        "PEwBasEPS": 1,
        "PEwAuditBascEPS": 2,
        "PEwDilutEPS": 3,
    }

    for suffix in PE_RATIO_SUFFIXES:
        marker = f"_{suffix}"
        if column.endswith(marker):
            date_value = parse_date_from_column(column)
            return (suffix_order.get(suffix, 99), -date_value.toordinal(), column)

    return (len(PE_RATIO_SUFFIXES), 0, column)


def is_shareholding_column(column):
    """Return True for flattened shareholding percentage columns."""
    return "Share Holding Percentage" in column


def order_columns_for_analysis(columns):
    """Return a finance-analysis-friendly column order without dropping data."""
    original_columns = list(columns)
    ordered_columns = []

    fixed_groups = [
        IDENTITY_COLUMNS,
        PRICE_COLUMNS,
        LIQUIDITY_COLUMNS,
        MARKET_VALUE_COLUMNS,
        SHARE_CAPITAL_COLUMNS,
        DEBT_AND_EQUITY_COLUMNS,
    ]

    for group in fixed_groups:
        ordered_columns.extend(existing_columns(original_columns, group))

    eps_columns = sorted(
        [column for column in original_columns if is_eps_metric_column(column)],
        key=eps_metric_sort_key,
    )
    ordered_columns.extend(eps_columns)

    audited_annual_columns = sorted(
        [column for column in original_columns if is_annual_performance_column(column)],
        key=annual_performance_sort_key,
    )
    ordered_columns.extend(audited_annual_columns)

    pe_columns = sorted(
        [column for column in original_columns if is_pe_ratio_column(column)],
        key=pe_ratio_sort_key,
    )
    ordered_columns.extend(pe_columns)

    ordered_columns.extend(existing_columns(original_columns, CORPORATE_ACTION_COLUMNS))

    shareholding_columns = [
        column for column in original_columns if is_shareholding_column(column)
    ]
    ordered_columns.extend(shareholding_columns)

    ordered_set = set(ordered_columns)
    remaining_columns = [
        column for column in original_columns if column not in ordered_set
    ]

    return ordered_columns + remaining_columns


def safe_divide(numerator, denominator):
    """Return numerator / denominator when both values are usable."""
    if pd.isna(numerator) or pd.isna(denominator) or denominator == 0:
        return pd.NA
    return numerator / denominator


def pct_change(new_value, old_value):
    """Return percentage change from old_value to new_value."""
    if pd.isna(new_value) or pd.isna(old_value) or old_value == 0:
        return pd.NA
    return ((new_value - old_value) / abs(old_value)) * 100


def clip_score(value):
    """Keep a scoring component between 0 and 100."""
    if pd.isna(value):
        return pd.NA
    return max(0, min(100, round(float(value), 2)))


def numeric_series(df, column):
    """Read a DataFrame column as numeric, returning NA when absent."""
    if column not in df.columns:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="Float64")
    return pd.to_numeric(df[column], errors="coerce")


def latest_annual_columns(df, prefix):
    """Return annual columns for a prefix, latest year first."""
    columns = [column for column in df.columns if column.startswith(f"{prefix}_")]
    return sorted(columns, key=annual_performance_sort_key)


def latest_valid_from_columns(df, columns):
    """Return each row's first non-empty value from an ordered column list."""
    if not columns:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="Float64")

    values = pd.DataFrame(
        {column: pd.to_numeric(df[column], errors="coerce") for column in columns},
        index=df.index,
    )
    return values.bfill(axis=1).iloc[:, 0]


def latest_valid_pe(df):
    """Choose the most useful current P/E value from available dynamic columns."""
    suffix_preference = ["PEwTrailRatio", "PEwBasEPS", "PEwAuditBascEPS"]
    all_candidates = []

    for suffix in suffix_preference:
        candidates = [column for column in df.columns if column.endswith(f"_{suffix}")]
        candidates = sorted(candidates, key=pe_ratio_sort_key)
        all_candidates.extend(candidates)

    return latest_valid_from_columns(df, all_candidates)


def latest_scraped_eps(df):
    """Choose an EPS value for screening from audited and interim fields."""
    candidates = latest_annual_columns(df, "Aud_EPS_COP_Basic")
    fallback_columns = [
        "Annual_EPS",
        "Annual_EPS_COP",
        "9Months_EPS",
        "9Months_EPS_COP",
        "Q3_EPS",
        "Q3_EPS_COP",
        "HalfYearly_EPS",
        "HalfYearly_EPS_COP",
        "Q2_EPS",
        "Q2_EPS_COP",
        "Q1_EPS",
        "Q1_EPS_COP",
    ]
    candidates.extend([column for column in fallback_columns if column in df.columns])
    return latest_valid_from_columns(df, candidates)


def annual_growth_from_prefix(df, prefix):
    """Calculate latest-to-oldest annual growth and available year count."""
    columns = latest_annual_columns(df, prefix)
    if len(columns) < 2:
        empty = pd.Series([pd.NA] * len(df), index=df.index, dtype="Float64")
        return empty, pd.Series([len(columns)] * len(df), index=df.index)

    values = pd.DataFrame(
        {column: pd.to_numeric(df[column], errors="coerce") for column in columns},
        index=df.index,
    )
    available_years = values.notna().sum(axis=1)
    latest = values.bfill(axis=1).iloc[:, 0]
    oldest = values[columns[::-1]].bfill(axis=1).iloc[:, 0]
    growth = pd.Series(
        [pct_change(new, old) for new, old in zip(latest, oldest)],
        index=df.index,
        dtype="Float64",
    )
    growth = growth.where(available_years >= 2, pd.NA)
    return growth, available_years


def trimmed_mean(series):
    """Return a simple 10%-90% trimmed mean for sector comparisons."""
    clean = pd.to_numeric(series, errors="coerce").dropna()
    if clean.empty:
        return pd.NA
    if len(clean) < 5:
        return round(clean.mean(), 2)

    lower = clean.quantile(0.10)
    upper = clean.quantile(0.90)
    trimmed = clean[(clean >= lower) & (clean <= upper)]
    return round(trimmed.mean(), 2) if not trimmed.empty else round(clean.mean(), 2)


def safe_round(value, digits=2):
    """Round regular numbers while preserving missing values."""
    if pd.isna(value):
        return pd.NA
    return round(float(value), digits)


def first_matching_shareholding_value(row, labels):
    """Extract a named holding percentage from flattened shareholding text."""
    shareholding_columns = [
        column for column in row.index if is_shareholding_column(column)
    ]
    values = []

    for column in shareholding_columns:
        text = str(row.get(column) or "")
        for label in labels:
            pattern = rf"{re.escape(label)}\s*:?\s*([0-9]+(?:\.[0-9]+)?)"
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                values.append(float(match.group(1)))

    return values[0] if values else pd.NA


def classify_liquidity(day_value):
    """Turn daily traded value into a practical liquidity label."""
    if pd.isna(day_value):
        return "Unknown"
    if day_value >= 50:
        return "High Liquidity"
    if day_value >= 10:
        return "Moderate Liquidity"
    if day_value >= 1:
        return "Low Liquidity"
    return "Illiquid"


def classify_debt_risk(debt_to_market_cap, debt_to_profit):
    """Turn debt ratios into a compact risk label."""
    if pd.isna(debt_to_market_cap) and pd.isna(debt_to_profit):
        return "Unknown"
    if (not pd.isna(debt_to_market_cap) and debt_to_market_cap >= 1) or (
        not pd.isna(debt_to_profit) and debt_to_profit >= 5
    ):
        return "High Debt Risk"
    if (not pd.isna(debt_to_market_cap) and debt_to_market_cap >= 0.4) or (
        not pd.isna(debt_to_profit) and debt_to_profit >= 3
    ):
        return "Moderate Debt Risk"
    return "Low Debt Risk"


def classify_dividend(dividend_yield):
    """Turn dividend yield into a compact shareholder-return label."""
    if pd.isna(dividend_yield):
        return "Unknown"
    if dividend_yield >= 6:
        return "Strong Dividend Support"
    if dividend_yield >= 3:
        return "Moderate Dividend Support"
    if dividend_yield > 0:
        return "Weak Dividend Support"
    return "No Dividend Support"


def format_text_list(title, items, fallback):
    """Format explanation lists for readable Excel cells."""
    unique_items = list(dict.fromkeys(items))
    if not unique_items:
        return f"{title}\n- {fallback}"
    return f"{title}\n" + "\n".join(f"- {item}" for item in unique_items)


def build_analysis_text(row):
    """Build signals, risks, conclusions, and next-check guidance for one stock."""
    positives = []
    negatives = []
    risks = []
    next_checks = []

    pe = row.get("Latest P/E Used")
    pe_vs_sector = row.get("P/E vs Sector Median (%)")
    price_to_nav = row.get("Price to NAV")
    eps = row.get("Latest EPS Used")
    profit = row.get("Latest Profit Used (mn)")
    cash_flow_to_profit = row.get("Cash Flow to Profit")
    debt_to_market_cap = row.get("Debt to Market Cap")
    debt_to_profit = row.get("Debt to Profit")
    dividend_yield = row.get("Dividend Yield %")
    price_position = row.get("Price Position 52W (%)")
    value_score = row.get("Value Score")
    quality_score = row.get("Quality Score")
    risk_score = row.get("Risk Score")
    final_score = row.get("Final Screening Score")
    reliability = row.get("Metric Reliability Score")
    liquidity = row.get("Liquidity Signal")
    outlier_flag = row.get("Outlier Flag")
    status = str(row.get("Present Operational Status") or "")
    sponsor_holding = row.get("Sponsor/Director Holding %")
    institute_holding = row.get("Institute Holding %")
    foreign_holding = row.get("Foreign Holding %")

    if not pd.isna(pe_vs_sector) and pe_vs_sector <= -20:
        positives.append("P/E is meaningfully below sector median")
        next_checks.append(
            "Verify whether the low P/E is caused by recurring earnings or a one-time profit spike"
        )
    elif not pd.isna(pe_vs_sector) and pe_vs_sector >= 30:
        negatives.append("P/E is materially above sector median")
        next_checks.append(
            "Check whether earnings growth, brand strength, or sector leadership justifies the premium P/E"
        )

    if not pd.isna(price_to_nav) and price_to_nav <= 1:
        positives.append("price is at or below NAVPS")
        next_checks.append(
            "Review asset quality and whether NAV is backed by productive assets"
        )
    elif not pd.isna(price_to_nav) and price_to_nav >= 3:
        negatives.append("price is high compared with NAVPS")
        next_checks.append(
            "Compare P/NAV with sector leaders and check if ROE supports the premium"
        )

    if not pd.isna(eps) and eps > 0:
        positives.append("EPS is positive")
    elif not pd.isna(eps):
        negatives.append("EPS is negative")
        risks.append("loss-making or weak earnings base")
        next_checks.append(
            "Read the latest quarterly report to understand the source of negative EPS"
        )
    else:
        risks.append("EPS data is missing")
        next_checks.append(
            "Manually verify EPS from the latest interim and audited financial statements"
        )

    if not pd.isna(profit) and profit > 0:
        positives.append("audited profit is positive")
    elif not pd.isna(profit):
        negatives.append("audited profit is negative")
        risks.append("negative audited profit")
        next_checks.append(
            "Check whether losses are recurring, one-off, or caused by finance cost or inventory pressure"
        )
    else:
        risks.append("audited profit data is missing")
        next_checks.append(
            "Verify audited profit before relying on valuation multiples"
        )

    if not pd.isna(cash_flow_to_profit) and cash_flow_to_profit >= 0.8:
        positives.append("operating cash flow supports reported profit")
    elif not pd.isna(cash_flow_to_profit) and cash_flow_to_profit < 0:
        negatives.append("operating cash flow is negative against profit")
        risks.append("weak cash conversion")
        next_checks.append(
            "Inspect receivables, inventory, working capital movement, and cash-flow notes"
        )
    elif not pd.isna(cash_flow_to_profit) and cash_flow_to_profit < 0.5:
        negatives.append("cash conversion is weak")
        risks.append("profit quality needs verification")
        next_checks.append(
            "Check whether profit is translating into operating cash flow"
        )

    if not pd.isna(debt_to_market_cap) and debt_to_market_cap <= 0.2:
        positives.append("debt burden is low relative to market cap")
    elif not pd.isna(debt_to_market_cap) and debt_to_market_cap >= 0.6:
        negatives.append("debt is high relative to market cap")
        risks.append("balance-sheet leverage risk")
        next_checks.append(
            "Review loan maturity, interest rate exposure, and finance-cost trend"
        )

    if not pd.isna(debt_to_profit) and debt_to_profit >= 4:
        risks.append("loan balance is high compared with profit")
        next_checks.append(
            "Check how many years of profit would be needed to cover total loans"
        )

    if not pd.isna(dividend_yield) and dividend_yield >= 5:
        positives.append("dividend yield is attractive")
        next_checks.append(
            "Confirm dividend payout sustainability from EPS, cash flow, and historical dividend policy"
        )
    elif not pd.isna(dividend_yield) and dividend_yield == 0:
        negatives.append("no dividend yield support")
        next_checks.append(
            "Check whether the company retained earnings for growth or skipped dividend due to weakness"
        )

    if not pd.isna(price_position) and price_position <= 25:
        positives.append("price is close to the lower part of its 52-week range")
        next_checks.append(
            "Check if the low price reflects temporary market pessimism or a structural business problem"
        )
    elif not pd.isna(price_position) and price_position >= 80:
        negatives.append("price is near its 52-week high")
        next_checks.append(
            "Consider margin of safety and whether waiting for a better entry price is sensible"
        )

    if liquidity in ["Low Liquidity", "Illiquid"]:
        risks.append("trading liquidity is weak")
        next_checks.append(
            "Check average traded value over multiple days before assuming an easy entry or exit"
        )

    if outlier_flag == "Yes":
        risks.append("one or more metrics look unusual versus normal screening bounds")
        next_checks.append(
            "Inspect outlier metrics manually before using sector comparison"
        )

    if status and status.lower() != "active":
        risks.append("operational status is not active")
        next_checks.append(
            "Verify operational, regulatory, or trading-status issues before any purchase consideration"
        )

    if not pd.isna(sponsor_holding) and sponsor_holding < 30:
        risks.append("sponsor/director holding is relatively low")
        next_checks.append(
            "Review sponsor/director holding trend and governance disclosures"
        )

    if not pd.isna(institute_holding) and institute_holding < 5:
        next_checks.append(
            "Check why institutional ownership is low and whether that reflects liquidity or quality concerns"
        )

    if not pd.isna(foreign_holding) and foreign_holding > 0:
        next_checks.append(
            "Review recent foreign holding trend for accumulation or exit signals"
        )

    if pd.isna(reliability) or reliability < 45:
        signal = "Insufficient Data"
        decision = "Insufficient Data"
        reason = (
            "The available scraped data is not reliable enough for a useful primary valuation conclusion. "
            "Important valuation, earnings, balance-sheet, or liquidity fields are missing or incomplete."
        )
    elif (not pd.isna(eps) and eps <= 0) or (not pd.isna(profit) and profit <= 0):
        signal = "Speculative / Risky"
        decision = "Avoid for Now"
        reason = (
            "Earnings or audited profit is weak. In this situation, valuation multiples such as P/E can be misleading, "
            "so the company needs deeper review before any purchase consideration."
        )
    elif (
        not pd.isna(value_score)
        and not pd.isna(quality_score)
        and not pd.isna(risk_score)
        and value_score >= 65
        and quality_score >= 55
        and risk_score <= 45
    ):
        signal = "Potentially Undervalued"
        decision = "Strong Candidate for Further Study"
        reason = (
            "The company appears attractive on primary valuation measures, and the available quality and risk checks "
            "are acceptable for a first-pass screening review."
        )
    elif (
        not pd.isna(value_score)
        and not pd.isna(risk_score)
        and value_score >= 55
        and risk_score >= 60
    ):
        signal = "Value Trap Risk"
        decision = "Cheap but Risky"
        reason = (
            "Some valuation metrics look cheap, but risk indicators weaken the case. The discount may be justified "
            "if debt, cash flow, earnings quality, liquidity, or governance concerns are serious."
        )
    elif (
        (not pd.isna(pe_vs_sector) and pe_vs_sector >= 40)
        or (not pd.isna(price_to_nav) and price_to_nav >= 3)
        or (not pd.isna(price_position) and price_position >= 85)
    ):
        signal = "Potentially Overvalued"
        decision = "Skip or Monitor"
        reason = (
            "The stock looks expensive on at least one primary valuation or price-position measure. A premium may still "
            "be justified, but only if growth, quality, and sector position are strong enough."
        )
    elif not pd.isna(final_score) and final_score >= 60:
        signal = "Fairly Valued / Worth Watching"
        decision = "Worth Watching"
        reason = (
            "The stock has a balanced mix of valuation, quality, dividend, liquidity, and risk signals. It does not show "
            "a strong bargain signal, but it is good enough to keep under review."
        )
    else:
        signal = "Fairly Valued"
        decision = "Monitor"
        reason = (
            "No strong undervaluation or overvaluation signal is visible from the primary screening metrics. The company "
            "can be monitored until price, earnings, dividend, or sector comparison becomes more attractive."
        )

    if decision == "Strong Candidate for Further Study":
        next_checks.append(
            "Read the latest annual report, quarterly report, PSI, and management discussion before any purchase decision"
        )
        next_checks.append(
            "Compare margins, revenue trend, and competitive position against direct sector peers"
        )
    elif decision in ["Worth Watching", "Monitor", "Skip or Monitor"]:
        next_checks.append(
            "Track the next quarterly EPS, dividend declaration, price trend, and sector median valuation"
        )

    positive_text = format_text_list(
        "Positive signals",
        positives,
        "No strong positive signal was detected from the available metrics.",
    )
    negative_text = format_text_list(
        "Negative signals",
        negatives,
        "No major negative signal was detected from the available metrics.",
    )
    risk_text = format_text_list(
        "Key risks",
        risks,
        "No major primary-screening risk was flagged.",
    )
    next_check_text = format_text_list(
        "What to check next",
        next_checks,
        "Perform normal due diligence: annual report, latest quarter, PSI disclosures, sector outlook, liquidity, governance, and peer comparison.",
    )
    watchlist_reason = (
        f"Conclusion\n{reason}\n\n{positive_text}\n\n{risk_text}"
    )

    return pd.Series(
        {
            "Positive Signals": positive_text,
            "Negative Signals": negative_text,
            "Key Risks": risk_text,
            "Primary Valuation Signal": signal,
            "Primary Valuation Reason": reason,
            "Watchlist Decision": decision,
            "Watchlist Reason": watchlist_reason,
            "What To Check Next": next_check_text,
        }
    )


def build_processed_analysis(raw_df):
    """Create the processed valuation and watchlist analysis worksheet."""
    df = raw_df.copy()

    processed = pd.DataFrame(index=df.index)
    for column in [
        "Company Name",
        "Trading Code",
        "Sector",
        "Market Category",
        "Present Operational Status",
        "LTP",
        "Market Cap (mn)",
        "Day Value (mn)",
        "Short-term Loan (mn)",
        "Long-term Loan (mn)",
        "Total Loan (mn)",
    ]:
        processed[column] = df[column] if column in df.columns else pd.NA

    processed["Latest EPS Used"] = latest_scraped_eps(df)
    processed["Latest NAVPS Used"] = latest_valid_from_columns(
        df, latest_annual_columns(df, "Aud_NAVPS")
    )
    processed["Latest Profit Used (mn)"] = latest_valid_from_columns(
        df, latest_annual_columns(df, "Aud_Profit_mn")
    )
    processed["Latest Operating Cash Flow Used (mn)"] = latest_valid_from_columns(
        df, latest_annual_columns(df, "Aud_PCO_mn")
    )
    processed["Latest P/E Used"] = latest_valid_pe(df)

    ltp = numeric_series(processed, "LTP")
    navps = numeric_series(processed, "Latest NAVPS Used")
    eps = numeric_series(processed, "Latest EPS Used")
    profit = numeric_series(processed, "Latest Profit Used (mn)")
    operating_cash_flow = numeric_series(
        processed, "Latest Operating Cash Flow Used (mn)"
    )
    market_cap = numeric_series(processed, "Market Cap (mn)")
    total_loan = numeric_series(processed, "Total Loan (mn)")
    day_value = numeric_series(processed, "Day Value (mn)")
    low_52w = numeric_series(df, "52W Low")
    high_52w = numeric_series(df, "52W High")

    processed["Price to NAV"] = [
        safe_divide(price, nav) for price, nav in zip(ltp, navps)
    ]
    processed["Earnings Yield (%)"] = [
        (
            safe_divide(earnings, price) * 100
            if not pd.isna(safe_divide(earnings, price))
            else pd.NA
        )
        for earnings, price in zip(eps, ltp)
    ]
    processed["Dividend Yield %"] = numeric_series(df, "Last Div Yield %")
    processed["Price Position 52W (%)"] = [
        (
            safe_divide(price - low, high - low) * 100
            if not pd.isna(safe_divide(price - low, high - low))
            else pd.NA
        )
        for price, low, high in zip(ltp, low_52w, high_52w)
    ]
    processed["Debt to Market Cap"] = [
        safe_divide(debt, cap) for debt, cap in zip(total_loan, market_cap)
    ]
    processed["Debt to Profit"] = [
        safe_divide(debt, earn) for debt, earn in zip(total_loan, profit)
    ]
    processed["Cash Flow to Profit"] = [
        safe_divide(cash_flow, earn)
        for cash_flow, earn in zip(operating_cash_flow, profit)
    ]

    eps_growth, eps_years = annual_growth_from_prefix(df, "Aud_EPS_COP_Basic")
    profit_growth, profit_years = annual_growth_from_prefix(df, "Aud_Profit_mn")
    navps_growth, navps_years = annual_growth_from_prefix(df, "Aud_NAVPS")
    processed["EPS Growth Available Years"] = eps_years
    processed["EPS Growth (%)"] = eps_growth
    processed["Profit Growth Available Years"] = profit_years
    processed["Profit Growth (%)"] = profit_growth
    processed["NAVPS Growth Available Years"] = navps_years
    processed["NAVPS Growth (%)"] = navps_growth

    processed["Sponsor/Director Holding %"] = df.apply(
        lambda row: first_matching_shareholding_value(
            row, ["Sponsor/Director", "Sponsor/Directer"]
        ),
        axis=1,
    )
    processed["Institute Holding %"] = df.apply(
        lambda row: first_matching_shareholding_value(row, ["Institute"]),
        axis=1,
    )
    processed["Foreign Holding %"] = df.apply(
        lambda row: first_matching_shareholding_value(row, ["Foreign"]),
        axis=1,
    )
    processed["Public Holding %"] = df.apply(
        lambda row: first_matching_shareholding_value(row, ["Public"]),
        axis=1,
    )

    processed["Sector Market Cap Rank"] = processed.groupby("Sector")[
        "Market Cap (mn)"
    ].rank(
        ascending=False,
        method="min",
    )
    processed["Sector Liquidity Rank"] = processed.groupby("Sector")[
        "Day Value (mn)"
    ].rank(
        ascending=False,
        method="min",
    )

    valid_pe = numeric_series(processed, "Latest P/E Used").between(
        0, 100, inclusive="neither"
    )
    valid_pnav = numeric_series(processed, "Price to NAV").between(
        0, 10, inclusive="neither"
    )
    valid_dividend = numeric_series(processed, "Dividend Yield %").between(
        0, 50, inclusive="both"
    )

    sector_stats = build_sector_summary(processed, valid_pe, valid_pnav, valid_dividend)
    processed = processed.merge(sector_stats, on="Sector", how="left")

    processed["P/E vs Sector Median (%)"] = [
        pct_change(pe, median)
        for pe, median in zip(
            processed["Latest P/E Used"], processed["Sector Median P/E"]
        )
    ]
    processed["P/NAV vs Sector Median (%)"] = [
        pct_change(pnav, median)
        for pnav, median in zip(
            processed["Price to NAV"], processed["Sector Median P/NAV"]
        )
    ]
    processed["Dividend Yield vs Sector Median (%)"] = [
        pct_change(dividend, median)
        for dividend, median in zip(
            processed["Dividend Yield %"], processed["Sector Median Dividend Yield %"]
        )
    ]

    processed["Liquidity Signal"] = processed["Day Value (mn)"].apply(
        classify_liquidity
    )
    processed["Debt Risk Signal"] = processed.apply(
        lambda row: classify_debt_risk(
            row["Debt to Market Cap"], row["Debt to Profit"]
        ),
        axis=1,
    )
    processed["Dividend Signal"] = processed["Dividend Yield %"].apply(
        classify_dividend
    )

    processed[["Outlier Flag", "Outlier Reason", "Usable For Sector Average?"]] = (
        processed.apply(
            build_outlier_assessment,
            axis=1,
        )
    )
    processed["Sector Outlier Count"] = processed.groupby("Sector")[
        "Outlier Flag"
    ].transform(lambda values: int((values == "Yes").sum()))
    processed["Metric Reliability Score"] = processed.apply(
        metric_reliability_score, axis=1
    )
    processed["Value Score"] = processed.apply(value_score, axis=1)
    processed["Quality Score"] = processed.apply(quality_score, axis=1)
    processed["Risk Score"] = processed.apply(risk_score, axis=1)
    processed["Liquidity Score"] = processed["Day Value (mn)"].apply(liquidity_score)
    processed["Dividend Score"] = processed.apply(dividend_score, axis=1)
    processed["Final Screening Score"] = processed.apply(final_screening_score, axis=1)

    text_columns = processed.apply(build_analysis_text, axis=1)
    processed = pd.concat([processed, text_columns], axis=1)

    for column in processed.columns:
        if column not in [
            "Company Name",
            "Trading Code",
            "Sector",
            "Market Category",
            "Present Operational Status",
            "Liquidity Signal",
            "Debt Risk Signal",
            "Dividend Signal",
            "Usable For Sector Average?",
            "Outlier Flag",
            "Outlier Reason",
            "Positive Signals",
            "Negative Signals",
            "Key Risks",
            "Primary Valuation Signal",
            "Primary Valuation Reason",
            "Watchlist Decision",
            "Watchlist Reason",
            "What To Check Next",
        ]:
            processed[column] = pd.to_numeric(processed[column], errors="coerce")

    processed = processed.reindex(
        columns=existing_columns(processed.columns, PROCESSED_BASE_COLUMNS)
    )
    processed = move_columns_to_front(
        processed, ["Company Name", "Trading Code", "Sector"]
    )
    numeric_columns = processed.select_dtypes(include=["number"]).columns
    processed[numeric_columns] = processed[numeric_columns].round(2)

    return processed


def build_sector_summary(processed, valid_pe, valid_pnav, valid_dividend):
    """Build sector-level averages, medians, and outlier counts."""
    rows = []
    for sector, group in processed.groupby("Sector", dropna=False):
        group_index = group.index
        pe_values = group.loc[valid_pe.loc[group_index], "Latest P/E Used"]
        pnav_values = group.loc[valid_pnav.loc[group_index], "Price to NAV"]
        dividend_values = group.loc[valid_dividend.loc[group_index], "Dividend Yield %"]

        outlier_count = int(len(group) - valid_pe.loc[group_index].sum())
        rows.append(
            {
                "Sector": sector,
                "Sector Company Count": len(group),
                "Sector Valid P/E Count": int(valid_pe.loc[group_index].sum()),
                "Sector Avg P/E": safe_round(
                    pd.to_numeric(pe_values, errors="coerce").mean()
                ),
                "Sector Median P/E": safe_round(
                    pd.to_numeric(pe_values, errors="coerce").median()
                ),
                "Sector Trimmed Avg P/E": trimmed_mean(pe_values),
                "Sector Avg P/NAV": safe_round(
                    pd.to_numeric(pnav_values, errors="coerce").mean()
                ),
                "Sector Median P/NAV": safe_round(
                    pd.to_numeric(pnav_values, errors="coerce").median()
                ),
                "Sector Trimmed Avg P/NAV": trimmed_mean(pnav_values),
                "Sector Avg Dividend Yield %": safe_round(
                    pd.to_numeric(dividend_values, errors="coerce").mean()
                ),
                "Sector Median Dividend Yield %": safe_round(
                    pd.to_numeric(dividend_values, errors="coerce").median()
                ),
                "Sector Trimmed Avg Dividend Yield %": trimmed_mean(dividend_values),
                "Sector Median Market Cap (mn)": safe_round(
                    group["Market Cap (mn)"].median()
                ),
                "Sector Median Day Value (mn)": safe_round(
                    group["Day Value (mn)"].median()
                ),
                "Sector Outlier Count": outlier_count,
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=[
                "Sector",
                "Sector Company Count",
                "Sector Valid P/E Count",
                "Sector Avg P/E",
                "Sector Median P/E",
                "Sector Trimmed Avg P/E",
                "Sector Avg P/NAV",
                "Sector Median P/NAV",
                "Sector Trimmed Avg P/NAV",
                "Sector Avg Dividend Yield %",
                "Sector Median Dividend Yield %",
                "Sector Trimmed Avg Dividend Yield %",
                "Sector Median Market Cap (mn)",
                "Sector Median Day Value (mn)",
                "Sector Outlier Count",
            ]
        )

    return pd.DataFrame(rows)


def build_sector_summary_report(processed_df):
    """Create a sector worksheet with valuation stats and decision counts."""
    if processed_df.empty:
        return pd.DataFrame()

    valid_pe = pd.to_numeric(processed_df["Latest P/E Used"], errors="coerce").between(
        0, 100, inclusive="neither"
    )
    valid_pnav = pd.to_numeric(processed_df["Price to NAV"], errors="coerce").between(
        0, 10, inclusive="neither"
    )
    valid_dividend = pd.to_numeric(
        processed_df["Dividend Yield %"], errors="coerce"
    ).between(0, 50, inclusive="both")

    summary = build_sector_summary(processed_df, valid_pe, valid_pnav, valid_dividend)
    decision_counts = (
        processed_df.pivot_table(
            index="Sector",
            columns="Watchlist Decision",
            values="Trading Code",
            aggfunc="count",
            fill_value=0,
        )
        .reset_index()
        .rename_axis(None, axis=1)
    )
    actual_outliers = (
        processed_df.assign(Is_Outlier=processed_df["Outlier Flag"].eq("Yes"))
        .groupby("Sector", dropna=False)["Is_Outlier"]
        .sum()
        .reset_index(name="Actual Outlier Count")
    )

    summary = summary.merge(actual_outliers, on="Sector", how="left")
    summary = summary.merge(decision_counts, on="Sector", how="left")
    summary["Actual Outlier Count"] = (
        summary["Actual Outlier Count"].fillna(0).astype(int)
    )
    return summary


def build_outlier_assessment(row):
    """Flag unusual values that can distort sector averages or conclusions."""
    reasons = []

    pe = row.get("Latest P/E Used")
    pnav = row.get("Price to NAV")
    eps = row.get("Latest EPS Used")
    profit = row.get("Latest Profit Used (mn)")
    debt_to_market_cap = row.get("Debt to Market Cap")
    cash_flow_to_profit = row.get("Cash Flow to Profit")
    day_value = row.get("Day Value (mn)")

    if pd.isna(pe) or pe <= 0:
        reasons.append("P/E missing or not meaningful")
    elif pe > 100:
        reasons.append("P/E above normal screening bound")

    if not pd.isna(pnav) and pnav > 10:
        reasons.append("P/NAV above normal screening bound")

    if not pd.isna(eps) and eps <= 0:
        reasons.append("negative EPS")

    if not pd.isna(profit) and profit <= 0:
        reasons.append("negative audited profit")

    if not pd.isna(debt_to_market_cap) and debt_to_market_cap > 1:
        reasons.append("debt exceeds market cap")

    if not pd.isna(cash_flow_to_profit) and cash_flow_to_profit < 0:
        reasons.append("negative operating cash flow conversion")

    if pd.isna(day_value) or day_value < 1:
        reasons.append("very low trading value")

    outlier_flag = "Yes" if reasons else "No"
    usable = (
        "No"
        if any("P/E" in reason or "negative EPS" in reason for reason in reasons)
        else "Yes"
    )
    if outlier_flag == "Yes" and usable == "Yes":
        usable = "Limited"

    return pd.Series(
        {
            "Outlier Flag": outlier_flag,
            "Outlier Reason": "; ".join(reasons) if reasons else "None",
            "Usable For Sector Average?": usable,
        }
    )


def metric_reliability_score(row):
    """Score availability of core data used in the screening model."""
    checks = [
        "LTP",
        "Market Cap (mn)",
        "Latest EPS Used",
        "Latest NAVPS Used",
        "Latest Profit Used (mn)",
        "Latest P/E Used",
        "Day Value (mn)",
        "Dividend Yield %",
        "Total Loan (mn)",
        "Sector Median P/E",
    ]
    available = sum(not pd.isna(row.get(column)) for column in checks)
    return round((available / len(checks)) * 100, 2)


def value_score(row):
    """Score valuation cheapness from P/E, P/NAV, yield, and price position."""
    score = 0
    pe = row.get("Latest P/E Used")
    pe_vs_sector = row.get("P/E vs Sector Median (%)")
    pnav = row.get("Price to NAV")
    earnings_yield = row.get("Earnings Yield (%)")
    dividend_yield = row.get("Dividend Yield %")
    price_position = row.get("Price Position 52W (%)")

    if not pd.isna(pe_vs_sector):
        if pe_vs_sector <= -30:
            score += 25
        elif pe_vs_sector <= -10:
            score += 18
        elif pe_vs_sector <= 10:
            score += 10
        elif pe_vs_sector <= 30:
            score += 4
    elif not pd.isna(pe):
        if 0 < pe <= 10:
            score += 22
        elif pe <= 15:
            score += 15
        elif pe <= 25:
            score += 8

    if not pd.isna(pnav):
        if pnav <= 1:
            score += 22
        elif pnav <= 1.5:
            score += 16
        elif pnav <= 2.5:
            score += 9

    if not pd.isna(earnings_yield):
        if earnings_yield >= 10:
            score += 16
        elif earnings_yield >= 6:
            score += 10
        elif earnings_yield > 0:
            score += 5

    if not pd.isna(dividend_yield):
        if dividend_yield >= 6:
            score += 14
        elif dividend_yield >= 3:
            score += 9
        elif dividend_yield > 0:
            score += 3

    if not pd.isna(price_position):
        if price_position <= 25:
            score += 14
        elif price_position <= 55:
            score += 8
        elif price_position <= 75:
            score += 3

    return clip_score(score)


def quality_score(row):
    """Score primary earnings, cash-flow, growth, and status quality."""
    score = 0
    eps = row.get("Latest EPS Used")
    profit = row.get("Latest Profit Used (mn)")
    cash_flow_to_profit = row.get("Cash Flow to Profit")
    eps_growth = row.get("EPS Growth (%)")
    profit_growth = row.get("Profit Growth (%)")
    navps_growth = row.get("NAVPS Growth (%)")
    status = str(row.get("Present Operational Status") or "").lower()

    if not pd.isna(eps) and eps > 0:
        score += 20
    if not pd.isna(profit) and profit > 0:
        score += 20
    if not pd.isna(cash_flow_to_profit):
        if cash_flow_to_profit >= 1:
            score += 20
        elif cash_flow_to_profit >= 0.5:
            score += 12
        elif cash_flow_to_profit > 0:
            score += 5
    if not pd.isna(eps_growth) and eps_growth > 0:
        score += 12
    if not pd.isna(profit_growth) and profit_growth > 0:
        score += 12
    if not pd.isna(navps_growth) and navps_growth > 0:
        score += 10
    if status == "active":
        score += 6

    return clip_score(score)


def risk_score(row):
    """Score primary risk, where higher is riskier."""
    score = 0
    eps = row.get("Latest EPS Used")
    profit = row.get("Latest Profit Used (mn)")
    debt_to_market_cap = row.get("Debt to Market Cap")
    debt_to_profit = row.get("Debt to Profit")
    cash_flow_to_profit = row.get("Cash Flow to Profit")
    day_value = row.get("Day Value (mn)")
    price_position = row.get("Price Position 52W (%)")
    pnav = row.get("Price to NAV")
    pe = row.get("Latest P/E Used")
    outlier_flag = row.get("Outlier Flag")

    if pd.isna(eps) or eps <= 0:
        score += 18
    if pd.isna(profit) or profit <= 0:
        score += 18
    if not pd.isna(debt_to_market_cap):
        if debt_to_market_cap >= 1:
            score += 20
        elif debt_to_market_cap >= 0.5:
            score += 12
    if not pd.isna(debt_to_profit) and debt_to_profit >= 4:
        score += 10
    if not pd.isna(cash_flow_to_profit) and cash_flow_to_profit < 0.5:
        score += 12
    if pd.isna(day_value) or day_value < 1:
        score += 10
    elif day_value < 10:
        score += 5
    if not pd.isna(price_position) and price_position >= 85:
        score += 6
    if not pd.isna(pnav) and pnav >= 4:
        score += 8
    if not pd.isna(pe) and pe >= 40:
        score += 8
    if outlier_flag == "Yes":
        score += 8

    return clip_score(score)


def liquidity_score(day_value):
    """Score trading liquidity from daily traded value."""
    if pd.isna(day_value):
        return 0
    if day_value >= 100:
        return 100
    if day_value >= 50:
        return 85
    if day_value >= 10:
        return 65
    if day_value >= 1:
        return 35
    return 10


def dividend_score(row):
    """Score income support from dividend yield and profit availability."""
    dividend_yield = row.get("Dividend Yield %")
    profit = row.get("Latest Profit Used (mn)")
    cash_flow_to_profit = row.get("Cash Flow to Profit")
    score = 0

    if not pd.isna(dividend_yield):
        if dividend_yield >= 8:
            score += 60
        elif dividend_yield >= 5:
            score += 45
        elif dividend_yield >= 3:
            score += 30
        elif dividend_yield > 0:
            score += 15

    if not pd.isna(profit) and profit > 0:
        score += 20
    if not pd.isna(cash_flow_to_profit) and cash_flow_to_profit >= 0.5:
        score += 20

    return clip_score(score)


def final_screening_score(row):
    """Blend value, quality, dividend, liquidity, and risk into one score."""
    components = [
        row.get("Value Score"),
        row.get("Quality Score"),
        row.get("Dividend Score"),
        row.get("Liquidity Score"),
        row.get("Risk Score"),
    ]
    if any(pd.isna(component) for component in components):
        return pd.NA

    score = (
        row["Value Score"] * 0.35
        + row["Quality Score"] * 0.30
        + row["Dividend Score"] * 0.15
        + row["Liquidity Score"] * 0.10
        + (100 - row["Risk Score"]) * 0.10
    )
    return round(score, 2)


def build_watchlist(processed_df):
    """Create a focused sheet for names that deserve follow-up attention."""
    if processed_df.empty:
        return processed_df.copy()

    keep_decisions = [
        "Strong Candidate for Further Study",
        "Worth Watching",
        "Cheap but Risky",
        "Monitor",
    ]
    watchlist = processed_df[
        processed_df["Watchlist Decision"].isin(keep_decisions)
    ].copy()
    watchlist = watchlist.sort_values(
        by=["Final Screening Score", "Value Score", "Quality Score"],
        ascending=[False, False, False],
        na_position="last",
    )
    return move_columns_to_front(watchlist, ["Company Name", "Trading Code", "Sector"])


def build_data_quality_issues(processed_df):
    """Create a worksheet with rows that need manual data review."""
    if processed_df.empty:
        return processed_df.copy()

    mask = (
        (processed_df["Outlier Flag"] == "Yes")
        | (processed_df["Metric Reliability Score"] < 60)
        | (processed_df["Primary Valuation Signal"] == "Insufficient Data")
    )
    columns = [
        "Company Name",
        "Trading Code",
        "Sector",
        "Metric Reliability Score",
        "Outlier Flag",
        "Outlier Reason",
        "Usable For Sector Average?",
        "What To Check Next",
    ]
    return processed_df.loc[
        mask, existing_columns(processed_df.columns, columns)
    ].copy()


def build_workbook_guide():
    """Create a readable guide sheet explaining workbook logic and labels."""
    return pd.DataFrame(GUIDE_ROWS)


def autosize_worksheet_columns(writer, sheet_name, df):
    """Set practical worksheet column widths for readability."""
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "D2" if sheet_name in ["Processed_Analysis", "Watchlist"] else "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="0F5B66")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="B7C9CC"))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    explanation_columns = {
        "Positive Signals",
        "Negative Signals",
        "Key Risks",
        "Primary Valuation Reason",
        "Watchlist Reason",
        "What To Check Next",
        "Meaning",
        "How To Use",
        "Outlier Reason",
    }

    for index, column in enumerate(df.columns, start=1):
        sample_values = df[column].head(100).tolist()
        max_length = max(
            [len(str(column))] + [len(str(value)) for value in sample_values]
        )
        width_cap = 80 if column in explanation_columns else 55
        min_width = 28 if column in explanation_columns else 12
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=index).column_letter
        ].width = min(
            max(max_length + 2, min_width),
            width_cap,
        )

        if column in explanation_columns:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=index).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    if sheet_name in ["Processed_Analysis", "Watchlist", "Workbook_Guide"]:
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 72 if sheet_name != "Workbook_Guide" else 54


def export_company_rows_to_excel(company_rows):
    """Save raw and processed company analysis into one Excel workbook."""
    market_date = None

    if company_rows and isinstance(company_rows, list):
        market_date = company_rows[0].get("Market Date")

    if not market_date:
        market_date = "Unknown_Date"

    market_date = str(market_date).replace(",", "").replace(" ", "_").replace("/", "-")

    folder = "Export_Data"
    os.makedirs(folder, exist_ok=True)
    file_name = f"DSE_Data_{market_date}.xlsx"
    file_path = os.path.join(folder, file_name)

    raw_df = pd.DataFrame(company_rows)
    raw_df = raw_df.reindex(columns=order_columns_for_analysis(list(raw_df.columns)))

    processed_df = build_processed_analysis(raw_df)
    raw_export_df, dropped_empty_columns_df = split_all_empty_columns(raw_df)
    sector_summary_df = build_sector_summary_report(processed_df)
    watchlist_df = build_watchlist(processed_df)
    data_quality_df = build_data_quality_issues(processed_df)
    workbook_guide_df = build_workbook_guide()

    sheets = {
        "Workbook_Guide": workbook_guide_df,
        "Raw_Scraped_Data": raw_export_df,
        "Processed_Analysis": processed_df,
        "Sector_Summary": sector_summary_df,
        "Watchlist": watchlist_df,
        "Data_Quality_Issues": data_quality_df,
        "Dropped_Empty_Columns": dropped_empty_columns_df,
    }

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            autosize_worksheet_columns(writer, sheet_name, sheet_df)

    print(f"\nData saved to: {file_path}")
